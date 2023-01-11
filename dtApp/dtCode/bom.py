from flask import render_template, request, redirect, url_for, jsonify
from dtApp import app
import os
from openpyxl import load_workbook


@app.route("/bom", methods=['GET', 'POST'])
def upload_file():
    '''
    This route loads the Bill of Materials Exel file and displays it on the webpage.

    Note that app.root_path is the absolute path to the root directory containing the app code.
    This should be ok for developement but may need changing to app.instance_path or os.path.dirname(app.instance_path) for implementation

    '''
    filename = os.path.join(app.root_path, 'static', 'BoM.xlsx')
    datafile = load_workbook(filename)
    # datafile.sheetnames
    datafile.active = 0
    sheets = []
    for n in range(1, len(datafile.worksheets)+1):
        sheets.append(datafile.active)
        datafile.active = n

    return render_template('bom.html', sheets=sheets, sheetnames=datafile.sheetnames)
